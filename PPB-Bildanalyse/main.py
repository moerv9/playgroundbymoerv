import os
import random
import shutil
import openpyxl
import cv2

folPath = "./Arbeitsordner"
folPathZaehlung1= "./Arbeitsordner/1. Zählung/"
folPathZaehlung2= "./Arbeitsordner/2. Zählung/"
resultPath = "/Users/marvinottersberg/Documents/Python/randomPicsFromStudy/test/gezaehlteValidierungsbilder/"
schongeschaetzteBilderPath= "/Users/marvinottersberg/Documents/Python/randomPicsFromStudy/test/30Bilder/"
gezaehlteBilder = "/Users/marvinottersberg/Documents/Python/randomPicsFromStudy/test/Mittelwert/"
team = ['Henry', 'Henrike', 'Kristin','Hauke','Jona','Marvin']



def list_files(dir):
    r = []
    subdirs = [x[0] for x in os.walk(dir)]
    for subdir in subdirs:
        files = os.walk(subdir).__next__()[2] # [0] returns root, [1] returns all dir, [2] returns all the files
        if (len(files) > 0):
            for file in files:
                if file.endswith(".jpg") or file.endswith(".JPG") or file.endswith(".jpeg"): #file.endswith(".HEIC") or file.endswith(".heic")or
                    r.append(os.path.join(subdir, file))
    return r


def list_files_With_Path(dir):
    r = []
    pathList = []
    subdirs = [x[0] for x in os.walk(dir)]
    for subdir in subdirs:
        files = os.walk(subdir).__next__()[2] # [0] returns root, [1] returns all dir, [2] returns all the files
        if (len(files) > 0):
            for file in files:
                if file.endswith(".jpg") or file.endswith(".JPG") or file.endswith(".jpeg"): #file.endswith(".HEIC") or file.endswith(".heic")or
                    r.append(os.path.join(subdir, file))
                    pathList.append(file)
    return pathList

def printlist(list):
    for item in list:
        print(item)

def getRandomPicturesFromList(list):
    r = random.choice(list)
    #print("Zufälliges Bild aus der Liste: {0}".format(r))
    return r

def getXrandomPicsFromList(amount,list):
    l = []
    for p in range(amount):
        pic = getRandomPicturesFromList(list)
        l.append(pic)
    #print("{0} Bilder aus der Liste: {1}".format(amount,getRandomPicturesFromList(finaleListe)))
    return l

def createTeamFoldersIn(path):
    for items in team:
        result = os.path.join(path,items)
        os.mkdir(result)

def copyRandomImgIntoSubFolder(amount,dest):
    dirInDestFolder = [x[0] for x in os.walk(dest)]
    for subdirs in dirInDestFolder[1:]:
        print("Subdirs: " +str(subdirs))
        newRandomImgList=getXrandomPicsFromList(amount,finaleListe)
        print("NewRandomImgList: "+str(newRandomImgList))
        for items in newRandomImgList:
            shutil.copy2(items,subdirs)

def copyRandomImgIntoSpecificFolder(amount,dest,folder_path):
    target=os.path.join(dest,folder_path)
    newRandomImgList=getXrandomPicsFromList(amount,finaleListe)
    for items in newRandomImgList:
        shutil.copy2(items,target)

def delDirs(folder):
    folders = [x[0] for x in os.walk(folder)]
    for dirs in folders[1:]:
        files = os.walk(dirs).__next__()[1]
        print(files)
        shutil.rmtree(dirs)

#Liest die Dateien in der Reihenfolge aus und fügt sie einer tabelle hinzu
#Wichtig, damit in die Excel-Tabelle zuerst der Name der Person(also des Ordners) und dann alle Bilder kommen,
#sodass zugeordnet werden kann
def read_inOrder(dir):
    r=[]
    list_Img=[]
    subdirs = [x[0] for x in os.walk(dir)]
    for subdir in subdirs[1:]:
        files = os.walk(subdir).__next__()[2] # [0] returns root, [1] returns all dir, [2] returns all the files
        list_Img.append(subdir[18:])
        if (len(files) > 0):
            for file in files:
                if file.endswith(".jpg") or file.endswith(".JPG") or file.endswith(".HEIC") or file.endswith(".heic")or file.endswith(".jpeg"):
                    r.append(os.path.join(subdir, file))

    print(list_Img)
    return list_Img


def checkDuplicates(dir):
    filelist= list_files(dir)
    print("Liste für Duplikatcheck: ", filelist)
    if len(filelist) == len(set(filelist)):
        print("liste enthält keine duplikate")
    else:
        print("liste enthält duplikate")


#Schreibt in die Excel Datei alle Dateien und Ordner, die in der Liste übergeben werden
def rw_Img_to_Excel(list):
    wb = openpyxl.load_workbook("./test/30Bilder/excel30.xlsx")
    #sheets = wb.sheetnames
    template_sheet = wb['Schätzung']
    #r1c1 = template_sheet.cell(1,1).value

    #write in cell
    #template_sheet.cell(2,2,"test")

    row=2
    col=1
    for items in list:
        template_sheet.cell(row,col,items)
        row+=1
    wb.save("./test/30Bilder/excel30.xlsx")

def getImgListWithSameWidth(list):

    imgSizeList=[]


    #width: 2176,3672,2268,3024,1908,2976
    #3607 Bilder mit 2176 Pixel Width
    for items in list:
        img = cv2.imread(items)
        w, h, _ = img.shape
        #print('width: {0}, height: {1}'.format(w, h))
        if w == 2176:
            imgSizeList.append(items)
        if len(imgSizeList) > 50:
            break

    return imgSizeList








#Press the green button in the gutter to run the script.
if __name__ == '__main__':

    listZaehlung1 = list_files(folPathZaehlung1)
    listZaehlung2 = list_files(folPathZaehlung2)
    finaleListe = listZaehlung1 + listZaehlung2

    #print("\nAnzahl Bilder Zählung 1: {0}, Anzahl Bilder Zählung 2: {1}, Insgesamt haben wir {2} Bilder ".format(len(listZaehlung1),len(listZaehlung2),len(finaleListe)))
    #delDirs(resultPath)
    #createTeamFoldersIn(resultPath)
    #copyRandomImgIntoSubFolder(50,resultPath)
    #copy_tree("Ergebnis",workingDir)

    #copyRandomImgIntoSpecificFolder(9,resultPath,"AnacondaCount")
    #checkDuplicates(resultPath)

    #list=read_inOrder("./test/30Bilder/")
    #rw_Img_to_Excel(list)

    ##Gets 100 random img from list and then filters after same width of the img    #width: 2176,3672,2268,3024,1908,2976
    #randomItems = getXrandomPicsFromList(1000, finaleListe)
    #newList = getImgListWithSameWidth(randomItems)
    #for items in newList:
    #    shutil.copy2(items,resultPath)


    all30BilderList = list_files(gezaehlteBilder)
    sameWidthPics = getImgListWithSameWidth(all30BilderList)

    for items in sameWidthPics:
        print(items + "\n")

    random10 = getXrandomPicsFromList(10,sameWidthPics)
    for item in random10:
        shutil.copy2(item,resultPath)













